from openpyxl import load_workbook

def leer_datos_excel(ruta_archivo):
    try:
        workbook = load_workbook(filename=ruta_archivo)
        
        hoja = workbook.worksheets[0]
        
        datos = []
        
        for fila in hoja.iter_rows(values_only=True):
            datos.append(fila)
        
        return datos
    except FileNotFoundError:
        print(f"Error: El archivo '{ruta_archivo}' no fue encontrado.")
    except Exception as e:
        print(f"Ocurrió un error: {e}")

ruta = r"C:\Users\DIEGO\Desktop\datos_empleados.xlsx"

datos_extraidos = leer_datos_excel(ruta)

if datos_extraidos:
    for fila in datos_extraidos:
        print(fila)